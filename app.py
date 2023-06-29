import csv
import os
from csv import writer, DictReader
from email import generator
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import excel2img
import json
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border, Side, Alignment
from openpyxl.cell import Cell
import pandas as pd
import numpy as np
from copy import copy
from math import ceil
import base64

from database import select_content, convert_to_csv_structure
from flask import Flask, render_template, request

app = Flask(__name__)
global csv_path
global bad_suspects

@app.route("/", methods=('GET', 'POST'))
def index():
    global csv_path
    global bad_suspects

    csv_path = "data.csv"
    try: bad_suspects
    except:
        bad_suspects = bad_suspect()

    items = [{"field": "Lead Plant:", "placeholder": ""},
             {"field": "Status:", "placeholder": ""},
             {"field": "Subject:", "placeholder": ""},
             {"field": "Approved:", "placeholder": ""},
             {"field": "Created by:", "placeholder": "CDSID"},
             {"field": "Plant(s) Affected:",
              "placeholder": "List affected assembly plant(s), lead plant is to be listed first"},
             {"field": "Issue Date:", "placeholder": "DD-MMM-YY"},
             {"field": "Status Date:",
              "placeholder": 'DD-MMM-YY  Have the option to state here "No change since last report," if applicable.'},
             {"field": "As Approved By:", "placeholder": "Approver's name(s) and Organization(s)"},
             {"field": "BSAQ#:", "placeholder": "List BSAQ Number"},
             {"field": "WERS Alert#(s):", "placeholder": "List WERS Alert number (1 alert per supplier)"},
             {"field": "Global 8D or 14D #", "placeholder": "List global 8D or 14D number"},
             {"field": "Additional Vehicles Affected:", "placeholder": "MY (Program Name) (Program ID)"},
             {"field": "Plant(s) Clean Point Date(s):", "placeholder": "List each assembly plant's clean point dates"},
             {"field": "Plant(s) # of Units Affected:",
              "placeholder": "Provide the total number of vehicles affected for each vehicle line per plant.  This number will only change if the VIN list increases or decreases."},
             {"field": "Supplier Name/Part #(s):", "placeholder": "List supplier's name and associated part number(s)"},
             {"field": "Date Repair Parts Available:", "placeholder": "DD-MMM-YY"},
             {"field": "Root Cause Owner/Org:",
              "placeholder": "List Root Cause Owner’s name and their organization.  Choose from these organizations:  Design, Manufacturing, Supplier Design, Supplier Process"},
             {"field": "Repair Funding Source:",
              "placeholder": 'Choose from the following:  Launch, Program, Plant, Supplier QR (include QR#), Central Funding (H80), Other (please list type, not "other")'},
             {"field": "Date Repair Instructions Avail:", "placeholder": "DD-MMM-YY"},
             {"field": "Background/Concern Description:", "placeholder": "List information"},
             {"field": "Root Cause:", "placeholder": "List information"},
             {"field": "Interim Corrective Action (ICA):", "placeholder": "List information"},
             {"field": "Permanent Corrective Action (PCA):", "placeholder": "List information"},
             {"field": "Status/Next Steps:", "placeholder": "List information"},
             {"field": "New Issues/Updates:", "placeholder": "List information"},
             {"field": "Help Required:", "placeholder": "List information"},
             {"field": "Additional On-Site Support:", "placeholder": "List information"},
             {"field": "Other:", "placeholder": "List information"}]

    if request.method == 'POST':
        select_content(request)
        save_in_csv(request.form, request.files)
        send_email(request.form, request.files)

    plants = []
    subjects = []
    total_fields = []

    try:
        with open(csv_path, 'r', encoding='UTF8', newline='') as f:
            reader = csv.DictReader(f)
            for row in reader:
                value = []
                for key in row.keys():
                    value.append(row[key])
                    if key == "Lead Plant:" and row[key] not in plants:
                        plants.append(row[key])
                subjects.append(value.copy())

        with open(csv_path, 'r', encoding='UTF8', newline='') as f:
            reader = csv.DictReader(f)
            for row in reader:
                for key in row.keys():
                    total_fields.append(key)
                break
    except:
        pass
    
    a, b = convert_to_csv_structure()
    items_ = [items, a, b, plants, bad_suspects]

    return render_template("Index.html", items=items_)

def bad_suspect():
    # Open the current file source
    arq = '../../CLMLC.CSV'
    df_current = pd.read_csv (arq, header=0)
    df_total = df_current['Bad_Suspect_Count']
    
    result = {}
    for x, v in enumerate(df_current['StopShip Number']):
        teste = str(df_total[x])
        if teste != "nan":
            string_teste = df_total[x].replace("'", "\"").replace("(", "\"(").replace(")",")\"")
            result[v] = json.loads(string_teste)
        else:
            result[v] = "N/A"

    return result
    
def save_in_csv(item_fields, files):
    global csv_path

    result_dict = []
    result_names = []
    first_time = False
    must_add = False

    subjects = []
    csv_keys = []
    
    try:
        f = open(csv_path, "r")
    except:
        first_time = True

    if not first_time:
        with open(csv_path, 'r', encoding='UTF8', newline='') as f:
            reader = csv.DictReader(f)
            for row in reader:
                value = []
                for key in row.keys():
                    value.append(row[key])
                    if key not in csv_keys:
                        csv_keys.append(key)
                subjects.append(value.copy())
                
    for field in item_fields.keys():
        result_names.append(field)
    for field in files.keys():
        result_names.append(field)

    for k in result_names:
        if k not in csv_keys and len(csv_keys) > 0:
            must_add = True
            csv_keys.append(k)
            for s in subjects:
                s.append("N/A")
    
    if len(csv_keys) > 0:
        for field in csv_keys:
            try:
                result_dict.append(item_fields[field])
            except:
                result_dict.append("N/A")
    else:
        for field in item_fields.keys():
            result_dict.append(item_fields[field])
            
    if not must_add:
        with open(csv_path, 'a', encoding='UTF8', newline='') as f:
            writer_object = writer(f)

            if first_time:
                writer_object.writerow(result_names)

            writer_object.writerow(result_dict)
            f.close()
    else:
        with open(csv_path, 'w', encoding='UTF8', newline='') as f:
            writer_object = writer(f)
            
            writer_object.writerow(csv_keys)

            for s in subjects:
                writer_object.writerow(s)
            writer_object.writerow(result_dict)
            f.close()

def send_email(form, files):
    print_in_excel(form, files)

    a = excel2img.export_img("Files/results.xlsx", "Files/test.png", "sheet", None)

    strFrom = 'from@example.com'
    strTo = 'to@example.com'

    msgRoot = MIMEMultipart('related')
    msgRoot['Subject'] = form["Subject:"]
    msgRoot['From'] = strFrom
    msgRoot['To'] = strTo
    msgRoot.preamble = 'This is a multi-part message in MIME format.'

    msgAlternative = MIMEMultipart('alternative')
    msgRoot.attach(msgAlternative)

    msgText = MIMEText('This is the alternative plain text message.')
    msgAlternative.attach(msgText)

    msgText = MIMEText('<img src="cid:image1">', 'html')
    msgAlternative.attach(msgText)

    fp = open('Files/test.png', 'rb')
    msgImage = MIMEImage(fp.read())
    fp.close()

    msgImage.add_header('Content-ID', '<image1>')
    msgRoot.attach(msgImage)
    msgRoot.add_header('X-Unsent', '1')

    outfile_name = 'draft_stop_shipment.eml'
    with open(outfile_name, 'w') as outfile:
        gen = generator.Generator(outfile)
        gen.flatten(msgRoot)

    os.startfile(outfile_name)

def patternCell(cell, targetFill, text=None, border=None, font_size=None, font_weight=None, alignment=None, underline='none'):
    size = "10" if font_size is None else font_size
    bold = False if font_weight is None else font_weight
    name = "Arial"
    alignment = 'center' if alignment is None else alignment

    fontStyle = Font(size=size, bold=bold, name=name, underline=underline)
    
    try:
        cell.fill = PatternFill("solid", start_color=targetFill)
    except:
        cell.fill = copy(targetFill)
    cell.alignment = openpyxl.styles.Alignment(horizontal=alignment, vertical='center', wrap_text=True)
    if border is None:
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
    else:
        if border is not False:
            cell.border = border

    cell.font = fontStyle
    if text is not None:
        cell.value = text

factor_of_font_size_to_width = {
    # TODO: other sizes
    12: {
        "factor": 0.8,  # width / count of symbols at row
        "height": 16
    }
}

def get_height_for_row(sheet, row_number, font_size=12):
    font_params = factor_of_font_size_to_width[font_size]
    row_ = list(sheet.rows)
    row = row_[row_number - 1]
    height = font_params["height"]

    for cell in row:
        try:
            words_count_at_one_row = sheet.column_dimensions[cell.column_letter].width*5 / font_params["factor"]
            lines = ceil(len(str(cell.value)) / words_count_at_one_row)
            height = max(height, lines * font_params["height"])
        except:
            pass

    return height

def print_in_excel(item_fields, files):
    wb = openpyxl.load_workbook('Files/template.xlsx')
    sheet = wb["sheet"]

    result_dict = []
    csv_keys = []

    for field in item_fields.keys():
        if "image" not in field:
            if field not in csv_keys:
                csv_keys.append(field)
            result_dict.append(item_fields[field])
    
    table_values = []
    count = 0
    position = 0

    sheet.column_dimensions["B"].width = 25
    sheet.column_dimensions["C"].width = 25
    sheet.column_dimensions["D"].width = 25
    sheet.column_dimensions["E"].width = 25
    sheet.column_dimensions["F"].width = 25
    sheet.column_dimensions["G"].width = 25
    sheet.column_dimensions["H"].width = 25

    for x, value in enumerate(result_dict):
        if "suspect" in csv_keys[x] or "nok" in csv_keys[x] or "ok" in csv_keys[x] or "repaired" in csv_keys[x] or "dtg" in csv_keys[x] or "notes" in csv_keys[x]:
            table_values.append((csv_keys[x], value))
            continue

        if "Plant(s) Clean Point Date(s):_" not in csv_keys[x] and "Date Repair Parts Available:_" not in csv_keys[x] and "Date Repair Instructions Avail:_" not in csv_keys[x]:
            patternCell(sheet['B' + str(position + 1)], targetFill="FFFFFF", text=csv_keys[x], font_weight=True, font_size=12, border=False, alignment="left", underline="single")

        if "Plant(s) Affected" in csv_keys[x]:
            value_ = ""
            for plant in value.split(","):
                if plant == "":
                    continue
                value_ += plant + ", "
            value = value_[:-2]
        
        patternCell(sheet['C' + str(position + 1)], targetFill="FFFFFF", text=value, font_weight=False, font_size=12, border=False, alignment="left")
        sheet.merge_cells('C' + str(position + 1) + ':I' + str(position + 1))

        sheet.column_dimensions["B"].width = 35
        sheet.row_dimensions[position + 1].height = get_height_for_row(sheet, position + 1, font_size=12)
        
        current_size = 0
        current_position = 0
        found = False
        for key in files.keys():
            if csv_keys[x] in key:
                try:
                    found = True
                    img = openpyxl.drawing.image.Image(files[key])
                    img.anchor = chr(67 + current_position) + str(position + 2)

                    width = 10 * (img.width/64)
                    if sheet.column_dimensions[chr(67 + current_position)].width is not None and width > sheet.column_dimensions[chr(67 + current_position)].width:
                        sheet.column_dimensions[chr(67 + current_position)].width = 10 * (img.width/64)
                    elif sheet.column_dimensions[chr(67 + current_position)].width is None:
                        sheet.column_dimensions[chr(67 + current_position)].width = 10 * (img.width/64)

                    if current_size < img.height:
                        sheet.row_dimensions[position + 2].height = img.height
                        current_size = img.height
                    sheet.add_image(img)
                    current_position += 1
                except:
                    pass
        if found:
            position += 1
        position += 1
        current_size = 0
        current_position = 0
        count = position

    count += 3
    
    patternCell(sheet['A1'], "ffffff", " ", border=False)

    sheet.row_dimensions[x + 1].height = 50

    up_down_heavy = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thick'), bottom=Side(style='thick'))

    # C - F | B - H | B - F | B - D
    patternCell(sheet['C' + str(count)], "ffffff", "Awaiting\nCHECK")
    patternCell(sheet['D' + str(count)], "ffffff", "Awaiting\nREPAIR")
    patternCell(sheet['E' + str(count)], "ffffff", "Checked  OK\n + OK with Repair")
    patternCell(sheet['F' + str(count)], "ffffff", "# of NOK\n units\n ACTUALLY REPAIRED")

    patternCell(sheet['B' + str(count + 1)], "ffffff", "Repair Status", border=False, font_weight=True)
    patternCell(sheet['C' + str(count + 1)], "D9D9D9", "Suspect", font_weight=True, border=up_down_heavy)
    patternCell(sheet['D' + str(count + 1)], "D9D9D9", "NOK", font_weight=True, border=up_down_heavy)
    patternCell(sheet['E' + str(count + 1)], "D9D9D9", "OK", font_weight=True, border=up_down_heavy)
    patternCell(sheet['F' + str(count + 1)], "D9D9D9", "Repaired", font_weight=True, border=up_down_heavy)
    patternCell(sheet['G' + str(count + 1)], "D9D9D9", "DTG", font_weight=True, border=up_down_heavy)
    patternCell(sheet['H' + str(count + 1)], "D9D9D9", "Notes", font_weight=True, border=up_down_heavy)
    
    plants = []
    for value in table_values:
        if value[0].split("_")[0] not in plants:
            plants.append(value[0].split("_")[0]) 
    
    subtotal_suspect = 0
    subtotal_nok = 0
    subtotal_ok = 0
    subtotal_repaired = 0

    count_x = 0
    for x, plant in enumerate(plants):
        count_x = x
        sheet.row_dimensions[count + 2 + x].height = 30
        patternCell(sheet['B' + str(count + 2 + x)], "ffffff", plant, font_weight=True)
        aux_count = 0
        for value in table_values:
            if plant in value[0]:
                if "suspect" in value[0]:
                    try:
                        subtotal_suspect += int(value[1])
                    except:
                        subtotal_suspect += 0
                elif "nok" in value[0]:
                    try:
                        subtotal_nok += int(value[1])
                    except:
                        subtotal_nok += 0
                elif "ok" in value[0]:
                    try:
                        subtotal_ok += int(value[1])
                    except:
                        subtotal_ok += 0
                elif "repaired" in value[0]:
                    try:
                        subtotal_repaired += int(value[1])
                    except:
                        subtotal_repaired += 0
                patternCell(sheet[chr(67 + aux_count) + str(count + 2 + x)], "ffffff", value[1])
                aux_count += 1
            else:
                aux_count = 0

    patternCell(sheet['B' + str(count + 3 + count_x)], "ffffff", "Sub-Total")
    patternCell(sheet['C' + str(count + 3 + count_x)], "ffffff", str(subtotal_suspect))
    patternCell(sheet['D' + str(count + 3 + count_x)], "ffffff", str(subtotal_nok))
    patternCell(sheet['E' + str(count + 3 + count_x)], "ffffff", str(subtotal_ok))
    patternCell(sheet['F' + str(count + 3 + count_x)], "ffffff", str(subtotal_repaired))

    patternCell(sheet['B' + str(count + 4 + count_x)], "ffffff", "Total Units on Hold:")
    patternCell(sheet['C' + str(count + 4 + count_x)], "ffffff", str(subtotal_suspect + subtotal_nok), border=up_down_heavy)

    wb.save("Files/results.xlsx")
